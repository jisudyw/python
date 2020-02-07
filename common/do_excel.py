
import openpyxl
from common.requests_test import Request
from common import contains
from  configparser import RawConfigParser

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.result=None

class DoExcel:

    def __init__(self,file_name=None,sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name

        try:
            self.workbook= openpyxl.load_workbook(self.file_name)
        except Exception as e:
            print("错误",e)
            #raise e
    def get_conf(self,section,option):
        rawconfig = RawConfigParser()  #创建对象
        rawconfig.read(contains.conf_file, encoding="utf-8")  #打开配置文件
        config_value=rawconfig.get(section=section, option=option)   #用配置文件中的section和option取value值
        return config_value
    def get_data(self,conf_case):
         sheet = self.workbook[self.sheet_name]
         casess = []
         if conf_case =="all":
             max_row=sheet.max_row
             for i in range(2, max_row + 1):
                 case = Case()
                 case.case_id = sheet.cell(row=i, column=1).value
                 case.title = sheet.cell(row=i, column=2).value
                 case.url = sheet.cell(row=i, column=3).value
                 case.data = sheet.cell(row=i, column=4).value
                 case.method = sheet.cell(row=i, column=5).value
                 case.expected = sheet.cell(row=i, column=6).value
                 casess.append(case)
         else :
             for i in eval(conf_case):
                        i=i+1
                        case=Case()
                        case.case_id=sheet.cell(row=i,column=1).value
                        case.title=sheet.cell(row=i,column=2).value
                        case.url = sheet.cell(row=i,column=3).value
                        case.data = sheet.cell(row=i, column=4).value
                        case.method = sheet.cell(row=i, column=5).value
                        case.expected = sheet.cell(row=i, column=6).value
                        casess.append(case)
         print(conf_case)
         return casess
    def write_data(self,row,actual,result):
        self.workbook[self.sheet_name].cell(row,7).value=actual  #写入实际结果
        self.workbook[self.sheet_name].cell(row,8).value=result
        self.workbook.save(self.file_name)
if __name__ == '__main__':
    do_excel=DoExcel(contains.data_path, 'register')
    conf_case=do_excel.get_conf("case","row")
    re = do_excel.get_data(conf_case)  # 调用常量，需要先导常量所在文件包，在进行 文件名.常量，调用常量路径
    print(re)
    for casess in re:
        print(casess.data)
        qq=Request().request(casess.method,casess.url,casess.data)
        if qq == casess.expected:  # 写入执行结果 PASS or FAIL
            do_excel.write_data(casess.case_id+1,qq,'PASS')
        else:
            do_excel.write_data(casess.case_id+1,qq,'FAIL')




